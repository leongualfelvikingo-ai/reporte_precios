import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import (
    getSampleStyleSheet
)


def exportar_excel(reporte, ruta):

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Criptomoneda"
    ws["B1"] = "Precio USD"
    ws["C1"] = "Fecha consulta"
    ws["D1"] = "Variación %"

    for i, item in enumerate(reporte, start=2):

        ws[f"A{i}"] = item["nombre"]
        ws[f"B{i}"] = item["precio"]
        ws[f"C{i}"] = item["fecha"]
        ws[f"D{i}"] = (
            item["variacion"]
            if item["variacion"] is not None
            else "N/A"
        )

    for cell in ["A1", "B1", "C1", "D1"]:

        ws[cell].font = Font(
            bold=True,
            color="FFFFFF"
        )

        ws[cell].fill = PatternFill(
            fill_type="solid",
            fgColor="4F81BD"
        )

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    wb.save(ruta)


def exportar_csv(reporte, ruta):

    with open(
        ruta,
        "w",
        newline="",
        encoding="utf-8"
    ) as archivo:

        writer = csv.writer(archivo)

        writer.writerow([
            "Criptomoneda",
            "Precio USD",
            "Fecha",
            "Variación %"
        ])

        for item in reporte:

            variacion = (
                item["variacion"]
                if item["variacion"] is not None
                else "N/A"
            )

            writer.writerow([
                item["nombre"],
                item["precio"],
                item["fecha"],
                variacion
            ])


def exportar_html(reporte, ruta):

    html = """
<html>
<head>
<meta charset="UTF-8">
<title>Reporte Cripto</title>
</head>
<body>

<h1>Reporte de Criptomonedas</h1>

<table border="1">
<tr>
<th>Criptomoneda</th>
<th>Precio USD</th>
<th>Fecha</th>
<th>Variación %</th>
</tr>
"""

    for item in reporte:

        variacion = (
            item["variacion"]
            if item["variacion"] is not None
            else "N/A"
        )

        html += f"""
<tr>
<td>{item['nombre']}</td>
<td>{item['precio']}</td>
<td>{item['fecha']}</td>
<td>{variacion}</td>
</tr>
"""

    html += """
</table>

</body>
</html>
"""

    with open(
        ruta,
        "w",
        encoding="utf-8"
    ) as archivo:

        archivo.write(html)

def exportar_pdf(reporte, ruta):

    pdf = SimpleDocTemplate(ruta)
    styles = getSampleStyleSheet()
    contenido = []

    contenido.append(
        Paragraph("Reporte de Criptomonedas", styles["Title"])
    )
    contenido.append(Spacer(1, 20))

    for item in reporte:

        variacion = item.get("variacion")
        var_str = f"{'+' if variacion > 0 else ''}{variacion}%" if variacion is not None else "N/A"

        contenido.append(Paragraph(f"<b>{item['nombre']}</b>", styles["Heading2"]))
        contenido.append(Paragraph(f"Precio: USD {item['precio']}", styles["Normal"]))
        contenido.append(Paragraph(f"Fecha: {item['fecha']}", styles["Normal"]))
        contenido.append(Paragraph(f"Variación: {var_str}", styles["Normal"]))
        contenido.append(Spacer(1, 12))

    pdf.build(contenido)