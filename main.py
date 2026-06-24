import requests
from openpyxl import Workbook
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from config import CRIPTOS, ARCHIVO_SALIDA
import csv

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import getSampleStyleSheet

# Crear carpeta si no existe
if not os.path.exists("output"):
    os.makedirs("output")

# Construir URL dinámica
ids = ",".join(CRIPTOS)
url = f"https://api.coingecko.com/api/v3/simple/price?ids={ids}&vs_currencies=usd"

# Obtener datos (con manejo de error básico)
try:
    response = requests.get(url)
    response.raise_for_status()
    data = response.json()
except Exception as e:
    print("Error al obtener datos:", e)
    exit()

# Fecha actual
ahora = datetime.now(
    ZoneInfo("America/Argentina/Buenos_Aires")
    ).strftime("%Y-%m-%d %H:%M")

# Crear estructura de datos reutilizable
reporte = []

for cripto in CRIPTOS:
    precio = data.get(cripto, {}).get("usd", "No disponible")

    reporte.append({
        "nombre": cripto.capitalize(),
        "precio": precio,
        "fecha": ahora
    })

# =========================
# MOSTRAR EN CONSOLA
# =========================

print("\nREPORTE DE CRIPTOMONEDAS\n")

for item in reporte:
    print(
        f"{item['nombre']}: "
        f"USD {item['precio']} "
        f"({item['fecha']})"
    )

# =========================
# GENERAR EXCEL
# =========================

wb = Workbook()
ws = wb.active

ws["A1"] = "Criptomoneda"
ws["B1"] = "Precio USD"
ws["C1"] = "Fecha consulta"

for i, item in enumerate(reporte, start=2):
    ws[f"A{i}"] = item["nombre"]
    ws[f"B{i}"] = item["precio"]
    ws[f"C{i}"] = item["fecha"]

wb.save(ARCHIVO_SALIDA)

# =========================
# GENERAR CSV
# =========================

with open(
    "output/reporte.csv",
    "w",
    newline="",
    encoding="utf-8"
) as archivo:

    writer = csv.writer(archivo)

    writer.writerow([
        "Criptomoneda",
        "Precio USD",
        "Fecha"
    ])

    for item in reporte:
        writer.writerow([
            item["nombre"],
            item["precio"],
            item["fecha"]
        ])

# =========================
# GENERAR HTML
# =========================

html = """
<html>
<head>
<meta charset="UTF-8">
<title>Reporte Cripto</title>
</head>
<body>

<h1>Reporte de Criptomonedas</h1>

<table border="1">
<tr>
<th>Criptomoneda</th>
<th>Precio USD</th>
<th>Fecha</th>
</tr>
"""

for item in reporte:
    html += f"""
<tr>
<td>{item['nombre']}</td>
<td>{item['precio']}</td>
<td>{item['fecha']}</td>
</tr>
"""

html += """
</table>

</body>
</html>
"""

with open(
    "output/reporte.html",
    "w",
    encoding="utf-8"
) as archivo:
    archivo.write(html)

# =========================
# GENERAR PDF
# =========================

pdf = SimpleDocTemplate(
    "output/reporte.pdf"
)

styles = getSampleStyleSheet()

contenido = []

contenido.append(
    Paragraph(
        "Reporte de Criptomonedas",
        styles["Title"]
    )
)

contenido.append(Spacer(1, 20))

for item in reporte:

    contenido.append(
        Paragraph(
            f"<b>{item['nombre']}</b>",
            styles["Heading2"]
        )
    )

    contenido.append(
        Paragraph(
            f"Precio: USD {item['precio']}",
            styles["Normal"]
        )
    )

    contenido.append(
        Paragraph(
            f"Fecha: {item['fecha']}",
            styles["Normal"]
        )
    )

    contenido.append(
        Spacer(1, 12)
    )

pdf.build(contenido)

# =========================
# RESUMEN FINAL
# =========================

print("\nArchivos generados correctamente:\n")

print(f"- Excel : {ARCHIVO_SALIDA}")
print("- CSV   : output/reporte.csv")
print("- HTML  : output/reporte.html")
print("- PDF   : output/reporte.pdf")
